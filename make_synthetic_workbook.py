"""Build a synthetic MAGAZZINO/REGISTRO workbook matching the documented contract,
so the engine can be tested without the real (sensitive) file."""
import openpyxl
from datetime import datetime

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "MAGAZZINO"

# header rows 1-2 (the real file uses row 2 as headers; data starts row 3)
headers = ["GIACENZA", "N. OPERAZIONE", "DATA DI CARICO", "QUANTITA'", "TIPOLOGIA",
           "CALIBRO", "MARCA ARMA", "MODELLO", "MATR. ARMA", "MATR. CANNA", "MATR. AGG.",
           "FORNITORE", "COSTO", "COSTO IMBALLO", "COSTO COMPL.", "DATA DDT", "DATA FATTURA",
           "PREZZO DI VENDITA", "V", "N. ATA", "DATA DI SCARICO", "N. FATTURA",
           "DATA FATTURA", "CLIENTE"]
ws.cell(1, 1).value = "MAGAZZINO BCW (SINTETICO PER TEST)"
for c, h in enumerate(headers, start=1):
    ws.cell(2, c).value = h

# a few existing purchase rows (rows 3..5), two with N.OPERAZIONE -> REGISTRO
samples = [
    # n_op, date, qty, tipologia, cal, marca, modello, matr, canna, agg, forn, costo, imb
    (101, datetime(2026, 1, 10), 1, "PISTOLA SEMIAUTOMATICA", "9x21", "BERETTA", "92X", "AA1001", "N/D", "N/D", "Fornitore Uno", 400, 10),
    (102, datetime(2026, 1, 12), 1, "REVOLVER", ".357", "S&W", "686", "AA1002", "N/D", "N/D", "Fornitore Uno", 600, 12),
    (None, datetime(2026, 1, 15), 1, "ACCESSORI", "", "VARI", "OTTICA", "N/D", "N/D", "N/D", "Fornitore Due", 90, 0),
]
r = 3
for s in samples:
    n_op, d, qty, tip, cal, marca, mod, matr, canna, agg, forn, costo, imb = s
    ws.cell(r, 1).value = f"=+D{r}-S{r}"
    if n_op:
        ws.cell(r, 2).value = n_op
    ws.cell(r, 3).value = d
    ws.cell(r, 4).value = qty
    ws.cell(r, 5).value = tip
    ws.cell(r, 6).value = cal
    ws.cell(r, 7).value = marca
    ws.cell(r, 8).value = mod
    ws.cell(r, 9).value = matr
    ws.cell(r, 10).value = canna
    ws.cell(r, 11).value = agg
    ws.cell(r, 12).value = forn
    ws.cell(r, 13).value = costo
    ws.cell(r, 14).value = imb
    ws.cell(r, 15).value = f"=+M{r}+N{r}"
    r += 1

# a reserved-N.OP slot (no date/fornitore) — must be ignored by find_last_data_row
ws.cell(20, 2).value = 999

# a summary/category row far below — must be ignored
ws.cell(30, 5).value = "TOTALE PISTOLE"

# REGISTRO sheet
reg = wb.create_sheet("REGISTRO")
reg.cell(1, 1).value = "REGISTRO OPERAZIONI"
reg_headers = ["N. OPERAZIONE", "DATA CARICO", "QUANTITA", "TIPOLOGIA", "CALIBRO",
               "MARCA", "MODELLO", "MATR ARMA", "MATR CANNA", "MATR AGG", "FORNITORE",
               "COSTO", "(riserva)", "DATA SCARICO", "CLIENTE", "TITOLO"]
for c, h in enumerate(reg_headers, start=1):
    reg.cell(3, c).value = h
# existing registro entries for n_op 101,102 (vlookup formulas)
for i, n_op in enumerate([101, 102]):
    rr = 4 + i
    reg.cell(rr, 1).value = n_op
    for offset in range(1, 12):
        reg.cell(rr, offset + 1).value = f"=+VLOOKUP($A{rr},MAGAZZINO!$B$2:$M$5000,{offset+1},FALSE)"

# other reference sheets
for name in ("INVENTARIO", "BASE", "PRINT"):
    wb.create_sheet(name)

wb.save("MAGAZZINO BCW V45.xlsx")
print("synthetic workbook written")
