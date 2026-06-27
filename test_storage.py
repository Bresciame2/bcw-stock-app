#!/usr/bin/env python3
"""
Storage connection self-test for the BCW register.
==================================================
Run this ONCE after configuring a backend (Cloudflare R2 or other) to prove the
app can read, write, lock, and find the workbook BEFORE colleagues rely on it.

It does NOT touch your real register data — it writes/reads/deletes a tiny
throwaway probe object and tests the lock on that probe, then (read-only)
confirms the real workbook object exists and downloads cleanly.

USAGE
  Locally, export the same values you'll put in Streamlit secrets, then run:

    export BCW_STORAGE=s3
    export BCW_S3_BUCKET=bcw-magazzino
    export BCW_S3_ENDPOINT="https://<ACCOUNT_ID>.r2.cloudflarestorage.com"
    export AWS_ACCESS_KEY_ID=...
    export AWS_SECRET_ACCESS_KEY=...
    export AWS_REGION=auto
    export BCW_WORKBOOK_NAME="MAGAZZINO BCW fixed.xlsx"
    python test_storage.py

Exit code 0 = all good. Non-zero = something to fix (message explains what).
"""
import sys
import storage


def ok(msg):   print(f"  \033[32m✓\033[0m {msg}")
def bad(msg):  print(f"  \033[31m✗\033[0m {msg}")


def main():
    print(f"Backend config: BCW_STORAGE={storage._cfg('BCW_STORAGE','local')!r}")
    try:
        backend = storage.get_backend()
    except Exception as e:
        bad(f"Could not build backend: {e}")
        return 2
    ok(f"backend created: {backend.name}")

    probe = "__bcw_selftest__.txt"
    payload = b"hello-bcw"

    # 1. write
    try:
        backend.write_bytes(probe, payload)
        ok("write_bytes")
    except Exception as e:
        bad(f"write failed (check bucket name, keys, endpoint, region): {e}")
        return 3

    # 2. read back + integrity
    try:
        got = backend.read_bytes(probe)
        assert got == payload
        ok("read_bytes (round-trip matches)")
    except Exception as e:
        bad(f"read failed or mismatch: {e}")
        return 4

    # 3. exists
    try:
        assert backend.exists(probe)
        ok("exists")
    except Exception as e:
        bad(f"exists failed: {e}")
        return 5

    # 4. lock acquire + release (this is what stops two colleagues corrupting it)
    try:
        backend.acquire_lock(probe, owner="selftest")
        ok("acquire_lock")
        backend.release_lock(probe)
        assert not backend.exists(probe + ".lock")
        ok("release_lock")
    except Exception as e:
        bad(f"lock cycle failed: {e}")
        return 6

    # 5. cleanup probe
    try:
        backend.delete(probe)
        assert not backend.exists(probe)
        ok("delete (cleanup)")
    except Exception as e:
        bad(f"delete failed: {e}")
        return 7

    # 6. real workbook present? (read-only; required before go-live)
    name = storage.workbook_name()
    if backend.exists(name):
        try:
            path = storage.download_workbook(backend)
            import openpyxl
            wb = openpyxl.load_workbook(path)
            sheets = wb.sheetnames
            ok(f"workbook '{name}' present and opens (sheets: {', '.join(sheets)})")
        except Exception as e:
            bad(f"workbook present but failed to open: {e}")
            return 8
    else:
        print(f"  \033[33m!\033[0m workbook '{name}' NOT in storage yet — "
              f"upload it once (R2 dashboard drag-drop, or the app's "
              f"Impostazioni tab). Everything else passed.")

    print("\nAll connection checks passed. ✅")
    return 0


if __name__ == "__main__":
    sys.exit(main())
