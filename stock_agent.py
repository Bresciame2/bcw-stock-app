"""
BCW Stock Management Agent
==========================
Called by Claude to update MAGAZZINO BCW V45.xlsx after reading purchase/sale documents.

Usage:
    python stock_agent.py carico  '<json>'   # add a purchase (DDT / fattura acquisto)
    python stock_agent.py scarico '<json>'   # record a sale (fattura vendita)
    python stock_agent.py cerca   '<json>'   # look up an item (by matr_arma or n_operazione)
    python stock_agent.py stato               # print current stock summary

CARICO JSON fields:
    data_carico      DD/MM/YYYY  – date goods arrived at Gardone (required)
    tipologia        str  – dropdown value (required)
    calibro          str
    marca            str
    modello          str
    matr_arma        str  – serial number of weapon (use "N/D" if none)
    matr_canna       str  – barrel serial (use "N/D" if none)
    matr_agg         str  – additional serial (use "N/D" if none)
    fornitore        str  – supplier name
    costo            float – purchase cost (ex-packaging)
    costo_imballo    float – packaging/shipping cost
    data_ddt         DD/MM/YYYY
    data_fattura     DD/MM/YYYY – purchase invoice date
    n_fattura        str  – purchase invoice number
    quantita         int  – default 1
    n_operazione     int  – REQUIRED only for weapons that enter REGISTRO
                            (progressive number, must be > current max)
    codice_fornitore str  – maker barcode scanned at intake -> stored in col Y
                            so a later scan maps back to this row
    genera_etichetta bool – if true, also produce a BCW QR+Code128 label PDF
    etichetta_out    str  – label PDF filename (default label_row<r>.pdf)

SCARICO JSON fields:
    code             str  – a scanned barcode string (resolved to the row)
    matr_arma        str  – find item by serial (preferred)
    n_operazione     int  – alternative lookup key
    prezzo_vendita   float
    data_scarico     DD/MM/YYYY
    n_fattura        str  – sale invoice number
    data_fattura_vendita  DD/MM/YYYY
    n_ata            str  – ATA number
    cliente          str  – buyer name
    titolo_acquisto  str  – title/notes for REGISTRO col P
    qty_venduta      int  – default 1
"""

import json
import sys
import os
import shutil
from datetime import datetime
import openpyxl
from openpyxl import load_workbook

# optional barcode integration — carico/scarico still work if libs are missing
try:
    import barcode_agent as _bc
    _BARCODE_OK = True
except Exception:
    _bc = None
    _BARCODE_OK = False

# ── paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(SCRIPT_DIR, "MAGAZZINO BCW V45.xlsx")

# ── helpers ──────────────────────────────────────────────────────────────────

def parse_date(s):
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s.strip(), fmt)
        except ValueError:
            pass
    return None


def load_wb():
    return load_workbook(EXCEL_FILE, data_only=False)


def save_wb(wb):
    # backup first
    backup = EXCEL_FILE.replace(".xlsx", "_backup.xlsx")
    if os.path.exists(EXCEL_FILE):
        shutil.copy2(EXCEL_FILE, backup)
    wb.save(EXCEL_FILE)


def find_last_data_row(ws):
    """Return the last PURCHASE row in MAGAZZINO.

    Uses ONLY col C (DATA DI CARICO / date) and col L (FORNITORE).
    Col B (N.OPERAZIONE) is excluded because rows 4601-4838 have N.OP
    pre-filled but no date/fornitore — those are reserved slots, not real
    purchases yet.  Col E (TIPOLOGIA) is excluded because summary/inventory
    rows (4852-4874) have category names there with no date or fornitore.
    """
    last = 2
    for r in range(3, ws.max_row + 1):
        c = ws.cell(r, 3).value   # DATA DI CARICO
        l = ws.cell(r, 12).value  # FORNITORE
        if any(v not in (None, "") for v in [c, l]):
            last = r
    return last


def get_max_operazione(ws):
    max_op = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, (int, float)) and v > max_op:
            max_op = int(v)
    return max_op


def find_item_row(ws, data):
    """
    Find row in MAGAZZINO for a given item.
    Searches by matr_arma (col I=9) first, then n_operazione (col B=2).
    Returns row number or None.
    """
    matr = str(data.get("matr_arma", "")).strip()
    n_op = data.get("n_operazione")

    for r in range(3, ws.max_row + 1):
        if matr and matr not in ("", "N/D"):
            cell_val = ws.cell(r, 9).value
            if cell_val is not None and str(cell_val).strip() == matr:
                return r
        if n_op:
            cell_val = ws.cell(r, 2).value
            if cell_val is not None and int(cell_val) == int(n_op):
                return r
    return None


# ── CARICO ───────────────────────────────────────────────────────────────────

def add_carico(data):
    wb = load_wb()
    ws = wb["MAGAZZINO"]

    # validate n_operazione uniqueness
    n_op = data.get("n_operazione")
    if n_op:
        n_op = int(n_op)
        max_op = get_max_operazione(ws)
        if n_op <= max_op:
            return {"status": "error",
                    "message": f"N. OPERAZIONE {n_op} già esiste o è minore del massimo ({max_op}). "
                               f"Usa un valore > {max_op}."}

    # find insert row (one after last filled)
    insert_row = find_last_data_row(ws) + 1

    r = insert_row

    # col A: GIACENZA formula
    ws.cell(r, 1).value = f"=+D{r}-S{r}"

    # col B: N. OPERAZIONE — preserve any pre-filled value (rows 4601-4838
    # already have N.OP numbers); only write if the cell is empty.
    existing_n_op = ws.cell(r, 2).value
    if existing_n_op not in (None, ""):
        n_op = int(existing_n_op)   # use the pre-filled value
    elif n_op:
        ws.cell(r, 2).value = n_op

    # col C: DATA DI CARICO
    ws.cell(r, 3).value = parse_date(data.get("data_carico"))

    # col D: QUANTITA'
    ws.cell(r, 4).value = int(data.get("quantita", 1))

    # col E-K: item fields
    ws.cell(r, 5).value  = data.get("tipologia", "")
    ws.cell(r, 6).value  = data.get("calibro", "")
    ws.cell(r, 7).value  = data.get("marca", "")
    ws.cell(r, 8).value  = data.get("modello", "")
    ws.cell(r, 9).value  = data.get("matr_arma", "N/D")
    ws.cell(r, 10).value = data.get("matr_canna", "N/D")
    ws.cell(r, 11).value = data.get("matr_agg", "N/D")

    # col L: FORNITORE
    ws.cell(r, 12).value = data.get("fornitore", "")

    # col M: COSTO
    costo = data.get("costo")
    if costo is not None:
        ws.cell(r, 13).value = float(costo)

    # col N: COSTO IMBALLO
    costo_imb = data.get("costo_imballo")
    if costo_imb is not None:
        ws.cell(r, 14).value = float(costo_imb)

    # col O: COSTO COMPL. formula
    ws.cell(r, 15).value = f"=+M{r}+N{r}"

    # col P: DATA DDT
    ws.cell(r, 16).value = parse_date(data.get("data_ddt"))

    # col Q: DATA FATTURA acquisto
    ws.cell(r, 17).value = parse_date(data.get("data_fattura"))

    # col Y: maker barcode scanned at intake (the "link" step)
    cod_forn = data.get("codice_fornitore")
    if cod_forn:
        ws.cell(r, 25).value = str(cod_forn).strip()

    save_wb(wb)

    # if weapon needs REGISTRO entry
    if n_op:
        wb2 = load_wb()
        _add_to_registro(wb2, n_op, r)
        save_wb(wb2)

    result = {
        "status": "success",
        "message": f"Carico aggiunto alla riga {r} del foglio MAGAZZINO.",
        "row": r,
        "n_operazione": n_op,
    }
    if cod_forn:
        result["codice_fornitore"] = str(cod_forn).strip()

    # optionally generate a BCW QR+Code128 label for this item
    if data.get("genera_etichetta"):
        if not _BARCODE_OK:
            result["etichetta_warning"] = "Librerie barcode non disponibili; etichetta non generata."
        else:
            gen = _bc.cmd_genera({
                "matr_arma": data.get("matr_arma"),
                "n_operazione": n_op,
                "out": data.get("etichetta_out", f"label_row{r}.pdf"),
            })
            if gen.get("status") == "success":
                result["etichetta_pdf"] = gen.get("pdf")
            else:
                result["etichetta_warning"] = gen.get("message")

    return result


def _add_to_registro(wb, n_operazione, mag_row):
    ws_reg = wb["REGISTRO"]

    # find next empty row in col A (after row 3)
    next_row = 4
    for r in range(4, ws_reg.max_row + 2):
        if ws_reg.cell(r, 1).value is None:
            next_row = r
            break

    # use a range that covers the new mag row + buffer
    upper = max(mag_row + 200, 5000)

    ws_reg.cell(next_row, 1).value = n_operazione

    # cols B–L: VLOOKUP offsets 2–12 from MAGAZZINO lookup table
    for offset in range(1, 12):
        ws_reg.cell(next_row, offset + 1).value = (
            f"=+VLOOKUP($A{next_row},"
            f"MAGAZZINO!$B$2:$M${upper},{offset + 1},FALSE)"
        )


# ── SCARICO ──────────────────────────────────────────────────────────────────

def add_scarico(data):
    wb = load_wb()
    ws = wb["MAGAZZINO"]

    # resolve by scanned barcode first, if provided
    row = None
    if data.get("code") and _BARCODE_OK:
        row, _matched = _bc.resolve_scan(ws, data["code"])
    if row is None:
        row = find_item_row(ws, data)
    if row is None:
        return {
            "status": "error",
            "message": "Articolo non trovato in MAGAZZINO. "
                       "Controlla code, matr_arma o n_operazione.",
        }

    # check not already sold
    if ws.cell(row, 19).value not in (None, 0, ""):
        existing_client = ws.cell(row, 24).value
        return {
            "status": "error",
            "message": f"Articolo alla riga {row} già venduto a: {existing_client}.",
        }

    # col R: PREZZO DI VENDITA
    pv = data.get("prezzo_vendita")
    if pv is not None:
        ws.cell(row, 18).value = float(pv)

    # col S: V (qty sold)
    ws.cell(row, 19).value = int(data.get("qty_venduta", 1))

    # col T: N. ATA
    if data.get("n_ata"):
        ws.cell(row, 20).value = str(data["n_ata"])

    # col U: DATA DI SCARICO
    ws.cell(row, 21).value = parse_date(data.get("data_scarico"))

    # col V: N. FATTURA vendita
    if data.get("n_fattura"):
        ws.cell(row, 22).value = str(data["n_fattura"])

    # col W: DATA FATTURA vendita
    ws.cell(row, 23).value = parse_date(data.get("data_fattura_vendita"))

    # col X: CLIENTE
    if data.get("cliente"):
        ws.cell(row, 24).value = data["cliente"]

    save_wb(wb)

    # update REGISTRO scarico cols if this weapon has N.OPERAZIONE
    n_op = ws.cell(row, 2).value
    if n_op:
        wb2 = load_wb()
        _update_registro_scarico(wb2, int(n_op), data)
        save_wb(wb2)

    return {
        "status": "success",
        "message": f"Scarico registrato alla riga {row}.",
        "row": row,
        "n_operazione": int(n_op) if n_op else None,
    }


def _update_registro_scarico(wb, n_operazione, data):
    ws_reg = wb["REGISTRO"]
    for r in range(4, ws_reg.max_row + 1):
        v = ws_reg.cell(r, 1).value
        if v is not None and int(v) == n_operazione:
            ws_reg.cell(r, 14).value = parse_date(data.get("data_scarico"))
            if data.get("cliente"):
                ws_reg.cell(r, 15).value = data["cliente"]
            if data.get("titolo_acquisto"):
                ws_reg.cell(r, 16).value = data["titolo_acquisto"]
            break


# ── CERCA ────────────────────────────────────────────────────────────────────

def cerca_item(data):
    wb = load_wb()
    ws = wb["MAGAZZINO"]

    row = find_item_row(ws, data)
    if row is None:
        return {"status": "not_found", "message": "Articolo non trovato."}

    headers = [
        "GIACENZA", "N_OPERAZIONE", "DATA_CARICO", "QUANTITA", "TIPOLOGIA",
        "CALIBRO", "MARCA", "MODELLO", "MATR_ARMA", "MATR_CANNA", "MATR_AGG",
        "FORNITORE", "COSTO", "COSTO_IMBALLO", "COSTO_COMPL",
        "DATA_DDT", "DATA_FATTURA_ACQ",
        "PREZZO_VENDITA", "V", "N_ATA", "DATA_SCARICO",
        "N_FATTURA_VEND", "DATA_FATTURA_VEND", "CLIENTE", "CODICE_A_BARRE",
    ]
    result = {}
    for col, header in enumerate(headers, start=1):
        val = ws.cell(row, col).value
        if isinstance(val, datetime):
            val = val.strftime("%d/%m/%Y")
        result[header] = val

    result["row"] = row
    result["status"] = "found"
    return result


# ── STATO ────────────────────────────────────────────────────────────────────

def stato():
    wb = load_wb()
    ws = wb["MAGAZZINO"]

    total = in_stock = sold = 0
    max_op = get_max_operazione(ws)
    last_row = find_last_data_row(ws)

    for r in range(3, ws.max_row + 1):
        tip = ws.cell(r, 5).value
        if tip in (None, ""):
            continue
        total += 1
        v = ws.cell(r, 19).value
        if v and int(v) > 0:
            sold += 1
        else:
            in_stock += 1

    return {
        "status": "ok",
        "totale_articoli": total,
        "in_giacenza": in_stock,
        "venduti": sold,
        "max_n_operazione": max_op,
        "ultimo_row_magazzino": last_row,
    }


# ── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"status": "error", "message": "Specifica: carico | scarico | cerca | stato"}))
        sys.exit(1)

    cmd = sys.argv[1].lower()

    if cmd == "stato":
        print(json.dumps(stato(), ensure_ascii=False, indent=2))
        sys.exit(0)

    if len(sys.argv) < 3:
        print(json.dumps({"status": "error", "message": "JSON dati mancante"}))
        sys.exit(1)

    try:
        data = json.loads(sys.argv[2])
    except json.JSONDecodeError as e:
        print(json.dumps({"status": "error", "message": f"JSON non valido: {e}"}))
        sys.exit(1)

    if cmd == "carico":
        result = add_carico(data)
    elif cmd == "scarico":
        result = add_scarico(data)
    elif cmd == "cerca":
        result = cerca_item(data)
    else:
        result = {"status": "error", "message": f"Comando sconosciuto: {cmd}"}

    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
