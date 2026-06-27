#!/usr/bin/env python3
"""
One-time: upload the local MAGAZZINO workbook into the configured shared
storage (Cloudflare R2), so the hosted app has its authoritative register.

Reads the same config as the app (env vars or Streamlit secrets). Safe to
re-run: it overwrites the object with whatever local file you point it at,
but it refuses to clobber an existing register unless you pass --force.

USAGE
    python3 seed_workbook.py                 # upload ./MAGAZZINO BCW fixed.xlsx
    python3 seed_workbook.py --force         # overwrite even if one exists
    python3 seed_workbook.py "/path/to.xlsx" # upload a specific file
"""
import os
import sys
import storage


def main():
    args = [a for a in sys.argv[1:] if a != "--force"]
    force = "--force" in sys.argv

    name = storage.workbook_name()
    local = args[0] if args else os.path.join(os.path.dirname(__file__), name)

    if not os.path.exists(local):
        print(f"  \033[31m✗\033[0m local file not found: {local}")
        return 2

    try:
        backend = storage.get_backend()
    except Exception as e:
        print(f"  \033[31m✗\033[0m could not build backend: {e}")
        return 3

    print(f"Backend : {backend.name}")
    print(f"Local   : {local}")
    print(f"Target  : {name}")

    if backend.exists(name) and not force:
        print(f"  \033[33m!\033[0m '{name}' ALREADY exists in storage. "
              f"Re-run with --force to overwrite. Nothing changed.")
        return 0

    try:
        storage.upload_workbook(backend, local)
    except Exception as e:
        print(f"  \033[31m✗\033[0m upload failed: {e}")
        return 4

    # verify it round-trips
    try:
        assert backend.exists(name)
        path = storage.download_workbook(backend)
        import openpyxl
        wb = openpyxl.load_workbook(path)
        print(f"  \033[32m✓\033[0m uploaded and verified — sheets: "
              f"{', '.join(wb.sheetnames)}")
    except Exception as e:
        print(f"  \033[31m✗\033[0m uploaded but verify failed: {e}")
        return 5

    print("\nWorkbook is now in shared storage. ✅")
    return 0


if __name__ == "__main__":
    sys.exit(main())
