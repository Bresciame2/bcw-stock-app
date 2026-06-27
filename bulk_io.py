"""
Bulk Excel templates + import.
==============================
Lets colleagues load many rows at once instead of typing them one by one:
download a ready-made .xlsx template for a given flow, fill it in, upload it.

Supported flows (`kind`):
    "carico"        -> new stock rows for the MAGAZZINO register
    "scarico"       -> sales rows to register
    "proforma"      -> proforma invoice line items (company-specific columns)
    "packing"       -> packing-list items grouped by parcel

Public API:
    template_bytes(kind, company=None) -> bytes        # the .xlsx to download
    parse(kind, file_bytes, company=None) -> list[dict] (or dict for packing)

Header matching is tolerant: case-insensitive, ignores spaces / underscores /
accents, so "Matr. Arma", "MATR_ARMA" and "matr arma" all map to the same key.
"""

import io
import unicodedata
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── column schemas ────────────────────────────────────────────────────────────
# Each schema: list of (Header shown in Excel, internal_key, example value)

CARICO = [
    ("DATA CARICO", "data_carico", "12/12/2025"),
    ("N OPERAZIONE", "n_operazione", 0),
    ("TIPOLOGIA", "tipologia", "FUCILE SEMIAUTOMATICO"),
    ("CALIBRO", "calibro", "12"),
    ("MARCA", "marca", "BENELLI"),
    ("MODELLO", "modello", "RAFFAELLO"),
    ("MATR ARMA", "matr_arma", "BS123456A"),
    ("MATR CANNA", "matr_canna", "CS123456E"),
    ("MATR AGG", "matr_agg", "N/D"),
    ("FORNITORE", "fornitore", "BENELLI ARMI SPA"),
    ("COSTO", "costo", 1000),
    ("COSTO IMBALLO", "costo_imballo", 0),
    ("DATA DDT", "data_ddt", "11/12/2025"),
    ("DATA FATTURA", "data_fattura", "11/12/2025"),
    ("QUANTITA", "quantita", 1),
]

SCARICO = [
    ("MATR ARMA", "matr_arma", "BS123456A"),
    ("CLIENTE", "cliente", "Armeria Rossi"),
    ("PREZZO VENDITA", "prezzo_vendita", 1200),
    ("DATA SCARICO", "data_scarico", "15/01/2026"),
    ("N FATTURA", "n_fattura", "123"),
    ("DATA FATTURA VENDITA", "data_fattura_vendita", "15/01/2026"),
    ("N ATA", "n_ata", ""),
    ("TITOLO ACQUISTO", "titolo_acquisto", ""),
]

PROFORMA_BME = [
    ("TYPE", "type", "SHOTGUN"),
    ("GAUGE", "gauge", "12"),
    ("BRAND", "brand", "BENELLI"),
    ("SERIAL", "serial", "BS123456A"),
    ("QTY", "qty", 1),
    ("PRICE", "price", 1200),
]

PROFORMA_BCW = [
    ("TIPO", "tipo", "FUCILE SEMIAUTOMATICO"),
    ("CAL", "cal", "12"),
    ("MARCA", "marca", "BENELLI"),
    ("DESC", "desc", "RAFFAELLO"),
    ("SERIAL", "serial", "BS123456A"),
    ("QTY", "qty", 1),
    ("PRICE", "price", 1200),
]

PACKING = [
    ("PARCEL", "parcel", 1),
    ("QTY", "qty", 1),
    ("TYPE", "type", "SHOTGUN"),
    ("BRAND", "brand", "BENELLI"),
    ("MODEL", "model", "RAFFAELLO"),
    ("CALIBER", "caliber", "12"),
    ("SERIAL1", "serial1", "BS123456A"),
    ("SERIAL2", "serial2", "CS123456E"),
    ("DIMS", "dims", "60x40x30 CM"),
    ("WEIGHT", "weight", "20,5 KG"),
]

_NUMERIC = {"costo", "costo_imballo", "quantita", "n_operazione",
            "prezzo_vendita", "qty", "price"}


def _schema(kind, company=None):
    if kind == "carico":
        return CARICO
    if kind == "scarico":
        return SCARICO
    if kind == "packing":
        return PACKING
    if kind == "proforma":
        return PROFORMA_BCW if (company == "BCW") else PROFORMA_BME
    raise ValueError(f"kind sconosciuto: {kind}")


def _norm(s):
    """Normalize a header for tolerant matching."""
    s = str(s or "")
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return "".join(ch for ch in s.lower() if ch.isalnum())


# ── template generation ───────────────────────────────────────────────────────

def template_bytes(kind, company=None):
    schema = _schema(kind, company)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DATI"

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    ex_font = Font(italic=True, color="808080")

    for c, (header, _key, example) in enumerate(schema, start=1):
        cell = ws.cell(1, c, header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        ws.cell(2, c, example).font = ex_font          # one greyed example row
        ws.column_dimensions[cell.column_letter].width = max(12, len(header) + 4)
    ws.freeze_panes = "A2"

    info = wb.create_sheet("ISTRUZIONI")
    notes = [
        "COME USARE QUESTO MODELLO",
        "",
        "1. Compila il foglio 'DATI', una riga per ogni articolo.",
        "2. La riga 2 (in grigio) è solo un esempio: sostituiscila o cancellala.",
        "3. Non rinominare le colonne dell'intestazione (riga 1).",
        "4. Le date vanno in formato GG/MM/AAAA (es. 12/12/2025).",
        "5. Salva il file e caricalo nell'app con il pulsante 'Carica Excel'.",
    ]
    if kind == "packing":
        notes.append("6. Colonna PARCEL: usa lo stesso numero per gli articoli "
                     "dello stesso collo. DIMS/WEIGHT si leggono dalla prima riga "
                     "di ciascun collo.")
    if kind == "carico":
        notes.append("6. N OPERAZIONE: 0 = non iscrivere a registro; "
                     "altrimenti un numero progressivo nuovo.")
    for i, line in enumerate(notes, start=1):
        cell = info.cell(i, 1, line)
        if i == 1:
            cell.font = Font(bold=True, size=13)
    info.column_dimensions["A"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── parsing ───────────────────────────────────────────────────────────────────

def _cell_to_str(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _coerce(key, v):
    if key in _NUMERIC:
        if v in (None, ""):
            return 0
        if isinstance(v, (int, float)):
            return v
        s = str(v).strip().replace("€", "").replace(" ", "")
        # European formatting: "1.234,50" -> dot=thousands, comma=decimal;
        # "1234,50" -> decimal comma; "1234.50" -> already a decimal point.
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        try:
            f = float(s)
            return int(f) if f.is_integer() else f
        except ValueError:
            return 0
    return _cell_to_str(v)


def _read_rows(file_bytes, schema):
    """Return list of dicts keyed by internal key, matching headers tolerantly."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["DATI"] if "DATI" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    by_norm = {_norm(h): key for h, key, _ in schema}
    header = rows[0]
    col_key = {}
    for idx, h in enumerate(header):
        k = by_norm.get(_norm(h))
        if k:
            col_key[idx] = k

    out = []
    for raw in rows[1:]:
        rec = {}
        for idx, key in col_key.items():
            val = raw[idx] if idx < len(raw) else None
            rec[key] = _coerce(key, val)
        if any(str(v).strip() for v in rec.values()):
            out.append(rec)
    return out


def parse(kind, file_bytes, company=None):
    """Parse an uploaded template.

    For carico/scarico/proforma -> list[dict].
    For packing -> {"n_parcels": int, "parcels": [{items,dims,weight}, ...]}.
    """
    schema = _schema(kind, company)
    rows = _read_rows(file_bytes, schema)

    if kind != "packing":
        # drop the greyed example row if the user left it verbatim
        return rows

    # group by parcel number, preserving first-seen order
    order = []
    groups = {}
    for r in rows:
        try:
            p = int(r.get("parcel") or 1)
        except (ValueError, TypeError):
            p = 1
        if p not in groups:
            groups[p] = {"items": [], "dims": "", "weight": ""}
            order.append(p)
        item = {k: r.get(k, "") for k in
                ("qty", "type", "brand", "model", "caliber", "serial1", "serial2")}
        if any(str(v).strip() for v in item.values()):
            groups[p]["items"].append(item)
        if not groups[p]["dims"] and str(r.get("dims") or "").strip():
            groups[p]["dims"] = str(r.get("dims")).strip()
        if not groups[p]["weight"] and str(r.get("weight") or "").strip():
            groups[p]["weight"] = str(r.get("weight")).strip()

    parcels = [groups[p] for p in order]
    return {"n_parcels": len(parcels), "parcels": parcels}
